#!/usr/bin/env python3
import boto3
import configparser
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Read AWS profiles from ~/.aws/config
aws_config_path = os.path.expanduser("~/.aws/config")
config = configparser.ConfigParser()
config.read(aws_config_path)

# Extract profiles (including "default" if it exists)
profiles = ["default"] if "default" in config.sections() else []
profiles += [section.replace("profile ", "") for section in config.sections() if section.startswith("profile ")]

# Dictionary to store DataFrames per profile
profile_dataframes = {}

for profile in profiles:
    print(f"Fetching instances for profile: {profile}")

    # Create a session with the profile
    session = boto3.Session(profile_name=profile)

    # Get available regions for EC2
    ec2_client = session.client("ec2")
    regions = [region["RegionName"] for region in ec2_client.describe_regions()["Regions"]]

    instances_data = []
    instance_type_counts = {}

    for region in regions:
        print(f"  Checking region: {region}")

        # Create a session in the region
        ec2 = session.client("ec2", region_name=region)

        # Fetch instances
        instances = ec2.describe_instances()

        # Get account ID
        sts_client = session.client("sts")
        account_id = sts_client.get_caller_identity()["Account"]

        for reservation in instances["Reservations"]:
            for instance in reservation["Instances"]:
                instance_type = instance["InstanceType"]
                state = instance["State"]["Name"]  # e.g., running, stopped

                # Extract instance name from tags (if exists)
                instance_name = next(
                    (tag["Value"] for tag in instance.get("Tags", []) if tag["Key"] == "Name"),
                    "N/A"
                )

                instances_data.append({
                    "Account ID": account_id,
                    "Region": region,
                    "Instance ID": instance["InstanceId"],
                    "Instance Type": instance_type,
                    "Instance Name": instance_name,
                    "State": state
                })

                # Count instance types
                if instance_type in instance_type_counts:
                    instance_type_counts[instance_type] += 1
                else:
                    instance_type_counts[instance_type] = 1

    # Create DataFrame for instances
    df_instances = pd.DataFrame(instances_data)

    # Create DataFrame for instance type counts
    df_counts = pd.DataFrame(list(instance_type_counts.items()), columns=["Instance Type", "Count"])

    # Store the data
    profile_dataframes[profile] = {"instances": df_instances, "counts": df_counts}

# Save to Excel with multiple sheets
output_file = "aws_instances_report.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for profile, data in profile_dataframes.items():
        data["instances"].to_excel(writer, sheet_name=f"{profile}_Instances", index=False)
        data["counts"].to_excel(writer, sheet_name=f"{profile}_Counts", index=False)

# Load workbook to adjust column width
wb = load_workbook(output_file)
for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Auto-adjust column width based on max content length
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

wb.save(output_file)

print(f"Report saved to {output_file} with auto-adjusted column width.")

