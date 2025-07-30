#!/usr/bin/env python3

import boto3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle

# Get all available AWS profiles from ~/.aws/config and ~/.aws/credentials
profiles = boto3.Session().available_profiles
output_file = "aws_reserved_instances.xlsx"
default_region = "eu-west-1"  # Default region if none is found in the profile

# Initialize an Excel writer
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Create an empty list to store all profiles' data
    all_data = []

    # Iterate over profiles and AWS regions
    for profile in profiles:
        session = boto3.Session(profile_name=profile)

        # Get the AWS Account ID using sts
        sts_client = session.client("sts")
        try:
            account_id = sts_client.get_caller_identity()["Account"]
        except Exception as e:
            print(f"Error fetching account ID for profile {profile}: {e}")
            continue
        
        # Get regions for each profile
        try:
            ec2_client = session.client("ec2")
            regions = [region["RegionName"] for region in ec2_client.describe_regions()["Regions"]]
        except Exception as e:
            print(f"Error fetching regions for profile {profile}: {e}")
            continue

        # Initialize an empty list to store data for the current profile
        data = []

        for region in regions:
            print(f"Fetching reserved instances for profile: {profile}, region: {region}")

            try:
                # Create EC2 client for the specific region
                ec2_client = session.client("ec2", region_name=region)

                # Get reserved instances
                response = ec2_client.describe_reserved_instances()

                # Get all running instances to find names
                ec2_instances = ec2_client.describe_instances(Filters=[{"Name": "instance-state-name", "Values": ["running"]}])

                # Map instance types to their corresponding names
                instance_name_map = {}
                for reservation in ec2_instances["Reservations"]:
                    for instance in reservation["Instances"]:
                        instance_type = instance["InstanceType"]
                        name_tag = next((tag["Value"] for tag in instance.get("Tags", []) if tag["Key"] == "Name"), "Unnamed")
                        if instance_type not in instance_name_map:
                            instance_name_map[instance_type] = []
                        instance_name_map[instance_type].append(name_tag)

                for ri in response["ReservedInstances"]:
                    instance_type = ri["InstanceType"]
                    matching_instances = instance_name_map.get(instance_type, ["No matching instance"])

                    # Extract tags as a formatted string
                    tags = "; ".join([f"{tag['Key']}:{tag['Value']}" for tag in ri.get("Tags", [])]) if "Tags" in ri else "No tags"

                    # Convert the 'Start' and 'End' to timezone-unaware datetimes
                    start = pd.to_datetime(ri["Start"]).tz_localize(None).strftime('%Y-%m-%d %H:%M:%S')
                    end = pd.to_datetime(ri["End"]).tz_localize(None).strftime('%Y-%m-%d %H:%M:%S')

                    # Append row to data list
                    data.append({
                        "Region": region,  # Ensure the region is included here
                        "InstanceType": instance_type,
                        "InstanceNames": ", ".join(matching_instances),
                        "AccountId": account_id,  # Use the fetched AWS Account ID
                        "ReservationId": ri["ReservedInstancesId"],
                        "State": ri["State"],
                        "Count": ri["InstanceCount"],
                        "Start": start,
                        "End": end,
                        "Tags": tags
                    })
            except Exception as e:
                print(f"Error fetching reserved instances for profile {profile}, region {region}: {e}")
                continue
        
        # Create a DataFrame for the current profile's data
        df = pd.DataFrame(data)
        
        # Write the data to a separate sheet for each profile
        df.to_excel(writer, sheet_name=profile, index=False)
        
        # Add this profile's data to the all_profiles list
        all_data.extend(data)
    
    # Create a DataFrame for all profiles combined
    all_profiles_df = pd.DataFrame(all_data)
    
    # Write the combined data to the "All Profiles" sheet
    all_profiles_df.to_excel(writer, sheet_name="All Profiles", index=False)

# Load the workbook to apply text wrapping and adjust column width
wb = load_workbook(output_file)

# Create a named style for text cells
text_style = NamedStyle(name="text_style", number_format="@")

# Iterate through all sheets to apply text wrapping and adjust column width
for sheet in wb.sheetnames:
    worksheet = wb[sheet]
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            # Apply the text style to all cells
            cell.style = text_style
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

# Save the workbook with applied formatting
wb.save(output_file)

print(f"Export completed: {output_file}")

